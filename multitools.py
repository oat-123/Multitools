import streamlit as st
import webbrowser
import subprocess
import pandas as pd
import random
from collections import defaultdict
from openpyxl.styles import Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import date
import io
import time
import gspread
from google.oauth2 import service_account

# ===== ระบบ Login =====
users = {
    "oat": {"password": "crma74", "sheet_name": "ชั้น4_พัน4"},
    "time": {"password": "crma74", "sheet_name": "ชั้น4_พัน1"},
    "chai": {"password": "crma74", "sheet_name": "ชั้น4_พัน3"}
}

st.sidebar.title("🔐 เข้าสู่ระบบ")
username = st.sidebar.text_input("ชื่อผู้ใช้")
password = st.sidebar.text_input("รหัสผ่าน", type="password")

if st.sidebar.button("เข้าสู่ระบบ"):
    if username in users and users[username]["password"] == password:
        st.session_state["logged_in"] = True
        st.session_state["username"] = username
        st.session_state["sheet_name"] = users[username]["sheet_name"]
        st.rerun()
    else:
        st.sidebar.error("ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง")

if not st.session_state.get("logged_in"):
    st.stop()

st.sidebar.success(f"ยินดีต้อนรับ {st.session_state['username']}")
st.sidebar.success(f"ฐานข้อมูล : {st.session_state['sheet_name']}")

def connect_gsheet(sheet_name: str):
    creds_dict = st.secrets["gcp_service_account"]
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    creds = service_account.Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    client = gspread.authorize(creds)
    sheet = client.open_by_url(
        "https://docs.google.com/spreadsheets/d/1PfZdCw2iL65CPTZzNsCnkhF7EVJNFZHRvYAXqeOJsSk/edit?usp=drivesdk"
    )
    worksheet = sheet.worksheet(sheet_name)
    return worksheet

worksheet = connect_gsheet(st.session_state["sheet_name"])

# --- ส่วนตกแต่งและ layout ---
def render_header():
    st.markdown("""
        <style>
        /* Dark theme background and text */
        body, .stApp, .main, .block-container {
            background: #181c24 !important;
            color: #f5f6fa !important;
        }
        .stSidebar, .css-1d391kg, .css-1lcbmhc, .css-1v0mbdj {
            background: #232733 !important;
        }
        .header-img-container {
            width: 100%;
            display: flex;
            justify-content: center;
            align-items: center;
            margin-bottom: 10px;
        }
        .header-img-container img {
            max-width: 220px;
            width: 100%;
            height: auto;
            display: block;
            margin: 0 auto;
            border-radius: 16px;
            box-shadow: 0 4px 24px rgba(0,0,0,0.25);
        }
        .title-text {
            font-size: 2.3rem;
            font-weight: bold;
            letter-spacing: 1px;
            color: #ff4b4b;
            text-shadow: 1px 2px 8px #00000055;
        }
        .subtitle-text {
            font-size: 1.2rem;
            color: #1f77b4;
            margin-bottom: 10px;
        }
        .stButton>button, .export-link {
            background: linear-gradient(90deg, #1f77b4 0%, #ff4b4b 100%);
            color: #fff;
            border: none;
            border-radius: 8px;
            font-size: 1.1rem;
            font-weight: 500;
            padding: 0.5em 1.2em;
            margin: 0.2em 0;
            box-shadow: 0 2px 8px rgba(0,0,0,0.10);
            transition: background 0.2s, color 0.2s;
        }
        .stButton>button:hover, .export-link:hover {
            background: linear-gradient(90deg, #ff4b4b 0%, #1f77b4 100%);
            color: #fff;
        }
        .card {
            background: #232733;
            border-radius: 14px;
            box-shadow: 0 2px 12px rgba(0,0,0,0.13);
            padding: 1.2em 0.5em 1em 0.5em;
            margin-bottom: 1em;
            text-align: center;
            border: 1px solid #22242c;
        }
        .stTextInput>div>div>input, .stTextArea textarea, .stNumberInput input, .stSelectbox>div>div>div>input, .stMultiSelect>div>div>div>input {
            background: #232733 !important;
            color: #f5f6fa !important;
            border-radius: 8px;
            border: 1px solid #353b48;
        }
        .stTextInput>div>div>input:focus, .stTextArea textarea:focus, .stNumberInput input:focus {
            border: 1.5px solid #1f77b4;
        }
        .stExpanderHeader {
            color: #1f77b4 !important;
        }
        .stDataFrame, .stTable {
            background: #232733 !important;
            color: #f5f6fa !important;
        }
        .st-bb, .st-cq, .st-cv, .st-cw, .st-cx, .st-cy, .st-cz, .st-da, .st-db, .st-dc, .st-dd, .st-de, .st-df, .st-dg, .st-dh, .st-di, .st-dj, .st-dk, .st-dl, .st-dm, .st-dn, .st-do, .st-dp, .st-dq, .st-dr, .st-ds, .st-dt, .st-du, .st-dv, .st-dw, .st-dx, .st-dy, .st-dz, .st-e0, .st-e1, .st-e2, .st-e3, .st-e4, .st-e5, .st-e6, .st-e7, .st-e8, .st-e9, .st-ea, .st-eb, .st-ec, .st-ed, .st-ee, .st-ef, .st-eg, .st-eh, .st-ei, .st-ej, .st-ek, .st-el, .st-em, .st-en, .st-eo, .st-ep, .st-eq, .st-er, .st-es, .st-et, .st-eu, .st-ev, .st-ew, .st-ex, .st-ey, .st-ez, .st-f0, .st-f1, .st-f2, .st-f3, .st-f4, .st-f5, .st-f6, .st-f7, .st-f8, .st-f9, .st-fa, .st-fb, .st-fc, .st-fd, .st-fe, .st-ff, .st-fg, .st-fh, .st-fi, .st-fj, .st-fk, .st-fl, .st-fm, .st-fn, .st-fo, .st-fp, .st-fq, .st-fr, .st-fs, .st-ft, .st-fu, .st-fv, .st-fw, .st-fx, .st-fy, .st-fz, .st-g0, .st-g1, .st-g2, .st-g3, .st-g4, .st-g5, .st-g6, .st-g7, .st-g8, .st-g9, .st-ga, .st-gb, .st-gc, .st-gd, .st-ge, .st-gf, .st-gg, .st-gh, .st-gi, .st-gj, .st-gk, .st-gl, .st-gm, .st-gn, .st-go, .st-gp, .st-gq, .st-gr, .st-gs, .st-gt, .st-gu, .st-gv, .st-gw, .st-gx, .st-gy, .st-gz, .st-h0, .st-h1, .st-h2, .st-h3, .st-h4, .st-h5, .st-h6, .st-h7, .st-h8, .st-h9, .st-ha, .st-hb, .st-hc, .st-hd, .st-he, .st-hf, .st-hg, .st-hh, .st-hi, .st-hj, .st-hk, .st-hl, .st-hm, .st-hn, .st-ho, .st-hp, .st-hq, .st-hr, .st-hs, .st-ht, .st-hu, .st-hv, .st-hw, .st-hx, .st-hy, .st-hz, .st-i0, .st-i1, .st-i2, .st-i3, .st-i4, .st-i5, .st-i6, .st-i7, .st-i8, .st-i9, .st-ia, .st-ib, .st-ic, .st-id, .st-ie, .st-if, .st-ig, .st-ih, .st-ii, .st-ij, .st-ik, .st-il, .st-im, .st-in, .st-io, .st-ip, .st-iq, .st-ir, .st-is, .st-it, .st-iu, .st-iv, .st-iw, .st-ix, .st-iy, .st-iz, .st-j0, .st-j1, .st-j2, .st-j3, .st-j4, .st-j5, .st-j6, .st-j7, .st-j8, .st-j9, .st-ja, .st-jb, .st-jc, .st-jd, .st-je, .st-jf, .st-jg, .st-jh, .st-ji, .st-jj, .st-jk, .st-jl, .st-jm, .st-jn, .st-jo, .st-jp, .st-jq, .st-jr, .st-js, .st-jt, .st-ju, .st-jv, .st-jw, .st-jx, .st-jy, .st-jz, .st-k0, .st-k1, .st-k2, .st-k3, .st-k4, .st-k5, .st-k6, .st-k7, .st-k8, .st-k9, .st-ka, .st-kb, .st-kc, .st-kd, .st-ke, .st-kf, .st-kg, .st-kh, .st-ki, .st-kj, .st-kk, .st-kl, .st-km, .st-kn, .st-ko, .st-kp, .st-kq, .st-kr, .st-ks, .st-kt, .st-ku, .st-kv, .st-kw, .st-kx, .st-ky, .st-kz, .st-l0, .st-l1, .st-l2, .st-l3, .st-l4, .st-l5, .st-l6, .st-l7, .st-l8, .st-l9, .st-la, .st-lb, .st-lc, .st-ld, .st-le, .st-lf, .st-lg, .st-lh, .st-li, .st-lj, .st-lk, .st-ll, .st-lm, .st-ln, .st-lo, .st-lp, .st-lq, .st-lr, .st-ls, .st-lt, .st-lu, .st-lv, .st-lw, .st-lx, .st-ly, .st-lz, .st-m0, .st-m1, .st-m2, .st-m3, .st-m4, .st-m5, .st-m6, .st-m7, .st-m8, .st-m9, .st-ma, .st-mb, .st-mc, .st-md, .st-me, .st-mf, .st-mg, .st-mh, .st-mi, .st-mj, .st-mk, .st-ml, .st-mm, .st-mn, .st-mo, .st-mp, .st-mq, .st-mr, .st-ms, .st-mt, .st-mu, .st-mv, .st-mw, .st-mx, .st-my, .st-mz, .st-n0, .st-n1, .st-n2, .st-n3, .st-n4, .st-n5, .st-n6, .st-n7, .st-n8, .st-n9, .st-na, .st-nb, .st-nc, .st-nd, .st-ne, .st-nf, .st-ng, .st-nh, .st-ni, .st-nj, .st-nk, .st-nl, .st-nm, .st-nn, .st-no, .st-np, .st-nq, .st-nr, .st-ns, .st-nt, .st-nu, .st-nv, .st-nw, .st-nx, .st-ny, .st-nz, .st-o0, .st-o1, .st-o2, .st-o3, .st-o4, .st-o5, .st-o6, .st-o7, .st-o8, .st-o9, .st-oa, .st-ob, .st-oc, .st-od, .st-oe, .st-of, .st-og, .st-oh, .st-oi, .st-oj, .st-ok, .st-ol, .st-om, .st-on, .st-oo, .st-op, .st-oq, .st-or, .st-os, .st-ot, .st-ou, .st-ov, .st-ow, .st-ox, .st-oy, .st-oz, .st-p0, .st-p1, .st-p2, .st-p3, .st-p4, .st-p5, .st-p6, .st-p7, .st-p8, .st-p9, .st-pa, .st-pb, .st-pc, .st-pd, .st-pe, .st-pf, .st-pg, .st-ph, .st-pi, .st-pj, .st-pk, .st-pl, .st-pm, .st-pn, .st-po, .st-pp, .st-pq, .st-pr, .st-ps, .st-pt, .st-pu, .st-pv, .st-pw, .st-px, .st-py, .st-pz, .st-q0, .st-q1, .st-q2, .st-q3, .st-q4, .st-q5, .st-q6, .st-q7, .st-q8, .st-q9, .st-qa, .st-qb, .st-qc, .st-qd, .st-qe, .st-qf, .st-qg, .st-qh, .st-qi, .st-qj, .st-qk, .st-ql, .st-qm, .st-qn, .st-qo, .st-qp, .st-qq, .st-qr, .st-qs, .st-qt, .st-qu, .st-qv, .st-qw, .st-qx, .st-qy, .st-qz, .st-r0, .st-r1, .st-r2, .st-r3, .st-r4, .st-r5, .st-r6, .st-r7, .st-r8, .st-r9, .st-ra, .st-rb, .st-rc, .st-rd, .st-re, .st-rf, .st-rg, .st-rh, .st-ri, .st-rj, .st-rk, .st-rl, .st-rm, .st-rn, .st-ro, .st-rp, .st-rq, .st-rr, .st-rs, .st-rt, .st-ru, .st-rv, .st-rw, .st-rx, .st-ry, .st-rz, .st-s0, .st-s1, .st-s2, .st-s3, .st-s4, .st-s5, .st-s6, .st-s7, .st-s8, .st-s9, .st-sa, .st-sb, .st-sc, .st-sd, .st-se, .st-sf, .st-sg, .st-sh, .st-si, .st-sj, .st-sk, .st-sl, .st-sm, .st-sn, .st-so, .st-sp, .st-sq, .st-sr, .st-ss, .st-st, .st-su, .st-sv, .st-sw, .st-sx, .st-sy, .st-sz, .st-t0, .st-t1, .st-t2, .st-t3, .st-t4, .st-t5, .st-t6, .st-t7, .st-t8, .st-t9, .st-ta, .st-tb, .st-tc, .st-td, .st-te, .st-tf, .st-tg, .st-th, .st-ti, .st-tj, .st-tk, .st-tl, .st-tm, .st-tn, .st-to, .st-tp, .st-tq, .st-tr, .st-ts, .st-tt, .st-tu, .st-tv, .st-tw, .st-tx, .st-ty, .st-tz, .st-u0, .st-u1, .st-u2, .st-u3, .st-u4, .st-u5, .st-u6, .st-u7, .st-u8, .st-u9, .st-ua, .st-ub, .st-uc, .st-ud, .st-ue, .st-uf, .st-ug, .st-uh, .st-ui, .st-uj, .st-uk, .st-ul, .st-um, .st-un, .st-uo, .st-up, .st-uq, .st-ur, .st-us, .st-ut, .st-uu, .st-uv, .st-uw, .st-ux, .st-uy, .st-uz, .st-v0, .st-v1, .st-v2, .st-v3, .st-v4, .st-v5, .st-v6, .st-v7, .st-v8, .st-v9, .st-va, .st-vb, .st-vc, .st-vd, .st-ve, .st-vf, .st-vg, .st-vh, .st-vi, .st-vj, .st-vk, .st-vl, .st-vm, .st-vn, .st-vo, .st-vp, .st-vq, .st-vr, .st-vs, .st-vt, .st-vu, .st-vv, .st-vw, .st-vx, .st-vy, .st-vz, .st-w0, .st-w1, .st-w2, .st-w3, .st-w4, .st-w5, .st-w6, .st-w7, .st-w8, .st-w9, .st-wa, .st-wb, .st-wc, .st-wd, .st-we, .st-wf, .st-wg, .st-wh, .st-wi, .st-wj, .st-wk, .st-wl, .st-wm, .st-wn, .st-wo, .st-wp, .st-wq, .st-wr, .st-ws, .st-wt, .st-wu, .st-wv, .st-ww, .st-wx, .st-wy, .st-wz, .st-x0, .st-x1, .st-x2, .st-x3, .st-x4, .st-x5, .st-x6, .st-x7, .st-x8, .st-x9, .st-xa, .st-xb, .st-xc, .st-xd, .st-xe, .st-xf, .st-xg, .st-xh, .st-xi, .st-xj, .st-xk, .st-xl, .st-xm, .st-xn, .st-xo, .st-xp, .st-xq, .st-xr, .st-xs, .st-xt, .st-xu, .st-xv, .st-xw, .st-xx, .st-xy, .st-xz, .st-y0, .st-y1, .st-y2, .st-y3, .st-y4, .st-y5, .st-y6, .st-y7, .st-y8, .st-y9, .st-ya, .st-yb, .st-yc, .st-yd, .st-ye, .st-yf, .st-yg, .st-yh, .st-yi, .st-yj, .st-yk, .st-yl, .st-ym, .st-yn, .st-yo, .st-yp, .st-yq, .st-yr, .st-ys, .st-yt, .st-yu, .st-yv, .st-yw, .st-yx, .st-yy, .st-yz, .st-z0, .st-z1, .st-z2, .st-z3, .st-z4, .st-z5, .st-z6, .st-z7, .st-z8, .st-z9, .st-za, .st-zb, .st-zc, .st-zd, .st-ze, .st-zf, .st-zg, .st-zh, .st-zi, .st-zj, .st-zk, .st-zl, .st-zm, .st-zn, .st-zo, .st-zp, .st-zq, .st-zr, .st-zs, .st-zt, .st-zu, .st-zv, .st-zw, .st-zx, .st-zy, .st-zz {
            background: #232733 !important;
            color: #f5f6fa !important;
        }
        @media (max-width: 600px) {
            .title-text {
                font-size: 1.3rem;
            }
            .subtitle-text {
                font-size: 1rem;
            }
            .header-img-container img {
                max-width: 120px;
            }
        }
        </style>
    """, unsafe_allow_html=True)
    st.markdown(
        '<div class="header-img-container">'
        '<img src="https://images4.alphacoders.com/112/1127690.png" alt="header" />'
        '</div>',
        unsafe_allow_html=True
    )
    st.markdown("""
        <div style='text-align: center;'>
            <div class='title-text'>
                <span style='color:#ff4b4b;'>J.A.R.V.I.S</span> <span style='color:#1f77b4;'>ระบบผู้ช่วย ฝอ.1</span>
            </div>
            <div class='subtitle-text'>⏬เลือกฟังก์ชันที่ต้องการจากด้านล่าง⏬</div>
            <hr style='border: 1px solid #ccc; margin-top: 10px; margin-bottom: 25px;'>
        </div>
    """, unsafe_allow_html=True)

def render_menu():
    st.markdown("""
        <style>
        .card {
            background: #f8f9fa;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.07);
            padding: 1.2em 0.5em 1em 0.5em;
            margin-bottom: 1em;
            text-align: center;
        }
        @media (max-width: 600px) {
            .stColumns {
                flex-direction: column !important;
            }
            .card {
                margin-bottom: 1.2em;
                padding: 1em 0.2em;
            }
        }
        .stButton>button {
            width: 100%;
            font-size: 1.1rem;
            border-radius: 8px;
        }
        </style>
    """, unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        st.markdown("#### 🛡 เวรรักษาการณ์")
        if st.button("ดู-อัพเดต", key="night_duty_btn"):
            st.session_state["mode"] = "night_duty"
        st.markdown("</div>", unsafe_allow_html=True)
    with col2:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        st.markdown("#### 📅 เวรเตรียมการ")
        if st.button("ดู-อัพเดต", key="weekend_duty_btn"):
            st.session_state["mode"] = "weekend_duty"
        st.markdown("</div>", unsafe_allow_html=True)
    with col3:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        st.markdown("#### 🏅 จัดยอดพิธี")
        if st.button("จัดยอด(สุ่ม)", key="ceremony_duty_btn"):
            st.session_state["mode"] = "ceremony_duty"
        st.markdown("</div>", unsafe_allow_html=True)
    col4, col5, _ = st.columns([1, 1, 1])
    with col4:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        st.markdown("#### 📝 ยอดปล่อย")
        if st.button("พิมพ์", key="home_btn"):
            st.session_state["mode"] = "home"
        st.markdown("</div>", unsafe_allow_html=True)
    with col5:
        st.markdown("<div class='card'>", unsafe_allow_html=True)
        st.markdown("#### 📊 สถิติโดนยอด")
        if st.button("อัพเดต-ตรวจสอบ", key="count_btn"):
            st.session_state["mode"] = "count"
        st.markdown("</div>", unsafe_allow_html=True)

def night_duty_mode():
    st.info("คุณเลือก: เวรยืนกลางคืน")
    # ลิงก์ CSV export (จากชีทแรก)
    csv_url = "https://docs.google.com/spreadsheets/d/1PjT38W2Zx7KV764yv9Vjwo9i0TJPacRI0iUGzP0ItAU/export?format=csv"

    try:
        # ให้ผู้ใช้เลือกระหว่างชีทแทกเวร หรือ ใบเวร (สรุป)
        sheet_option = st.radio(
            "เลือกดูชีท",
            ("แท็กเวร", "ใบเวร (สรุป)"))
        
        # สร้างลิงก์สำหรับชีทที่เลือก
        if sheet_option == "แท็กเวร":
            iframe_link = "https://docs.google.com/spreadsheets/d/e/2PACX-1vR8pO9068jsukCJL0guT_dF7I5cjYMMIhsu7ah-1DkPxSMxnYFsSkuRgffvSUJKVZzQccQyJEOPxvvg/pubhtml?gid=0&single=true&range=A1:I100"  # ลิงก์ชีทแทกเวร
            edit_link = "https://docs.google.com/spreadsheets/d/1PjT38W2Zx7KV764yv9Vjwo9i0TJPacRI0iUGzP0ItAU/edit#gid=0"  # ลิงก์แก้ไขชีทแทกเวร
        else:
            iframe_link = "https://docs.google.com/spreadsheets/d/e/2PACX-1vR8pO9068jsukCJL0guT_dF7I5cjYMMIhsu7ah-1DkPxSMxnYFsSkuRgffvSUJKVZzQccQyJEOPxvvg/pubhtml?gid=2030248910&single=true&range=A1:I100"  # ลิงก์ใบเวร (สรุป)
            edit_link = "https://docs.google.com/spreadsheets/d/1PjT38W2Zx7KV764yv9Vjwo9i0TJPacRI0iUGzP0ItAU/edit#gid=1"  # ลิงก์แก้ไขใบเวร (สรุป)
        
        # การฝัง iframe สำหรับชีทที่เลือก
        st.markdown(f"""
            <style>
                .iframe-container {{
                    width: 100%;
                    max-width: 1200px;
                    margin: auto;
                    border: 2px solid #4CAF50;
                    border-radius: 10px;
                    overflow: auto;
                    box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
                }}
        
                .iframe-container iframe {{
                    width: 100%;
                    min-width: 320px;
                    height: 1400px;
                    border: none;
                    transform: scale(1); 
                    transform-origin: top left;
                }}
        
                @media (max-width: 768px) {{
                    .iframe-container iframe {{
                        height: 900px;
                        min-width: 320px;
                        transform: scale(0.95);
                    }}
                }}
        
                .edit-link {{
                    text-align: right;
                    margin-top: 10px;
                    background-color: #4CAF50;
                    padding: 10px;
                    border-radius: 8px;
                    box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
                }}
                .edit-link a {{
                    color: white;
                    font-size: 16px;
                    text-decoration: none;
                    font-weight: bold;
                    transition: all 0.3s ease;
                }}
                .edit-link a:hover {{
                    color: #FFEB3B;
                    text-decoration: underline;
                }}
            </style>
        
            <div class="iframe-container">
                <iframe src="{iframe_link}"></iframe>
            </div>
        
            <div class="edit-link">
                <a href="{edit_link}" target="_blank">✏️ แก้ไข Google Sheets คลิกที่นี่</a>
            </div>
        """, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"โหลดข้อมูลไม่สำเร็จ: {e}")

def weekend_duty_mode():
    st.info("คุณเลือก: เวรเสาร์-อาทิตย์")
    st.markdown("[คลิกเพื่อดู Google Sheet📃](https://docs.google.com/spreadsheets/d/1ufm0LPa4c903jhlANKn_YqNyMtG9id0iN-tMHrhNRA8/edit?gid=1888956716)", unsafe_allow_html=True)

def home_mode():
    st.header("📋 พิมพ์ยอดปล่อย")

    # วันที่ของรายงาน
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("วันปล่อย", date.today())
    with col2:
        end_date = st.date_input("วันเข้ารร.", date.today())
    
    # ยอดเดิมแต่ละชั้นปี
    defaults = {5: 67, 4: 101, 3: 94, 2: 85}
    categories = ["เวรเตรียมพร้อม", "กักบริเวณ", "อยู่โรงเรียน","ราชการ", "โรงพยาบาล", "ลา", "อื่นๆ"]
    
    # กรอกข้อมูล
    st.subheader("กรอกข้อมูลเเต่ละชั้นปี")
    data = {}
    year_colors = {
        5: "#b3e5fc",  # สีฟ้าอ่อน
        4: "#c8e6c9",  # สีเขียวอ่อน
        3: "#bbdefb",  # สีน้ำเงินอ่อน
        2: "#f8bbd0",  # สีชมพูอ่อน
    }
    for year in [5, 4, 3, 2]:
        data[year] = {}
        with st.expander(f"ชั้นปีที่ {year}"):
            st.markdown(
                f"""
                <div style="background-color: {year_colors[year]}; padding: 15px; border-radius: 10px;">
                """,
                unsafe_allow_html=True
            )
            for cat in categories:
                val = st.number_input(
                    f"{cat} ชั้นปีที่ {year}",
                    min_value=0,
                    step=1,
                    key=f"{cat}_{year}"
                )
                data[year][cat] = val
            st.markdown("</div>", unsafe_allow_html=True)
    
    # ปุ่ม "สร้างรายงาน" และ "ทำไฟล์" ในแถวเดียวกัน ตกแต่งด้วย CSS
    col1, col2 = st.columns([1, 1])
    
    with col1:
        generate = st.button("📘 สร้างรายงาน")
    
    with col2:
        st.markdown("""
            <style>
            .stButton>button.export-link {
                background-color: #28a745;
                color: white;
                border-radius: 6px;
                padding: 0.4em 1em;
                border: none;
                font-size: 14px;
            }
            .stButton>button.export-link:hover {
                background-color: #218838;
                color: white;
            }
            </style>
            <a href="https://docs.google.com/spreadsheets/d/1_kKUegxtwwd3ce3EduPqRoPpgAF1_IcecA1ri9Pfxz0/edit?gid=351113778#gid=351113778" target="_blank">
                <button class="export-link">📗 ทำไฟล์</button>
            </a>
        """, unsafe_allow_html=True)
    
    # ✅ ทำงานต่อเมื่อกดปุ่ม
    if generate:
        st.success("รายงานถูกสร้างเรียบร้อย")
    
        lines = []
        start_str = start_date.strftime("%-d %b").replace("May", "พ.ค.").replace("Jun", "มิ.ย.")
        thai_year = end_date.year + 543  # แปลง ค.ศ. -> พ.ศ.
        end_str = end_date.strftime("%-d %b").replace("May", "พ.ค.").replace("Jun", "มิ.ย.") + f" {str(thai_year)[-2:]}"
    
        lines.append(f"พัน.4 กรม นนร.รอ. ขออนุญาตส่งยอด นนร. ปล่อยพักบ้าน, อยู่โรงเรียน และ เวรเตรียมพร้อม ของวันที่   {start_str} - {end_str} ดังนี้")
    
        for y in [5, 4, 3, 2]:
            lines.append(f"ชั้นปีที่ {y} ยอดเดิม {defaults[y]} นาย")
    
        def section(title, key):
            lines.append(f"{key+1}.{title}")
            total = 0
            for y in [5, 4, 3, 2]:
                val = data[y].get(title, 0)
                total += val
                show_val = f"{val}" if val != 0 else "-"
                lines.append(f"   -ชั้นปีที่ {y} จำนวน {show_val} นาย")
            show_total = f"{total}" if total != 0 else "-"
            lines.append(f"   -รวม {show_total} นาย")
    
        # 1. ยอดปล่อยบ้าน
        lines.append("1.ยอดปล่อยพักบ้าน")
        total_home = 0
        for y in [5, 4, 3, 2]:
            sum_others = sum(data[y].values())
            val = defaults[y] - sum_others
            total_home += val
            lines.append(f"   -ชั้นปีที่ {y} จำนวน {val} นาย")
        lines.append(f"   -รวม {total_home} นาย")
    
        for i, cat in enumerate(["อยู่โรงเรียน", "เวรเตรียมพร้อม", "กักบริเวณ", "โรงพยาบาล", "ราชการ", "ลา", "อื่นๆ"], start=2):
            section(cat, i)
        lines.append("จึงเรียนมาเพื่อกรุณาทราบ")
    
        st.text_area("รายงานยอด", value="\n".join(lines), height=600)

def count_mode():
 # STEP 1: ลิงก์ดูสถิติ
    sheet_id = "1PfZdCw2iL65CPTZzNsCnkhF7EVJNFZHRvYAXqeOJsSk"
    user_gid_map = {
        "oat": "0",
        "time": "589142731",
        "chai": "258225546",
    }

    users = {
        "oat": {"password": "crma74", "sheet_name": "ชั้น4_พัน4"},
        "time": {"password": "crma74", "sheet_name": "ชั้น4_พัน1"},
        "chai": {"password": "crma74", "sheet_name": "ชั้น4_พัน3"}
    }

    username = st.session_state.get("username", "")
    sheet_name = users.get(username, {}).get("sheet_name", username)
    gid = user_gid_map.get(username, "0")
    sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit#gid={gid}"

    st.markdown(f"🔍 [กดเพื่อดูสถิติโดนยอดปัจจุบัน (ชีท: {sheet_name})]({sheet_url})", unsafe_allow_html=True)

    # STEP 2: อัปโหลดไฟล์
    ยอด_file = st.file_uploader("📤 อัปโหลดไฟล์ยอด (.xlsx)", type="xlsx")

    if ยอด_file:
        try:
            xls = pd.ExcelFile(ยอด_file)
            sheet_names = xls.sheet_names

            selected_sheets = st.multiselect("📄 เลือกชีทที่ต้องการนับแต้ม", sheet_names)
            sheet_data = {}

            for sheet in selected_sheets:
                st.markdown(f"### 📌 ตั้งค่าความเหนื่อยสำหรับชีท: {sheet}")
                เหนื่อย = st.slider(f"ระดับความเหนื่อยของ '{sheet}' (1–5)", 1, 5, 3, key=sheet)

                try:
                    df = pd.read_excel(xls, sheet_name=sheet, header=None, skiprows=3)
                    df = df.dropna(how='all')

                    if df.shape[1] >= 4:
                        df["ชื่อเต็ม"] = df.iloc[:, 2].fillna("").astype(str).str.strip() + " " + df.iloc[:, 3].fillna("").astype(str).str.strip()
                        preview_df = pd.DataFrame({
                            "ลำดับ": df.iloc[:, 0],
                            "ชื่อ": df.iloc[:, 2],
                            "สกุล": df.iloc[:, 3],
                        })
                        st.dataframe(preview_df, use_container_width=True)

                        sheet_data[sheet] = {"df": df, "เหนื่อย": เหนื่อย}
                    else:
                        st.warning(f"⚠️ ไฟล์ชีท '{sheet}' มีคอลัมน์ไม่ครบ A–D")
                except Exception as e:
                    st.error(f"❌ ไม่สามารถอ่านชีท '{sheet}': {e}")

            if st.button("✅ อัปเดตแต้มเข้า Google Sheets"):
                try:
                    ws = connect_gsheet(sheet_name)
                    gsheet_data = ws.get_all_values()
                    gsheet_df = pd.DataFrame(gsheet_data)
                    gsheet_df.columns = gsheet_df.iloc[0]
                    gsheet_df = gsheet_df[1:].reset_index(drop=True)
            
                    gsheet_df["ชื่อเต็ม"] = gsheet_df.iloc[:, 2].astype(str).str.strip() + " " + gsheet_df.iloc[:, 3].astype(str).str.strip()
            
                    if "สถิติโดนยอด" not in gsheet_df.columns:
                        gsheet_df["สถิติโดนยอด"] = 0
                    gsheet_df["สถิติโดนยอด"] = pd.to_numeric(gsheet_df["สถิติโดนยอด"], errors='coerce').fillna(0).astype(int)
            
                    # ✅ รวมแต้มจากหลายชีทตามความเหนื่อย
                    for sheet, data in sheet_data.items():
                        df = data["df"]
                        เหนื่อย = data["เหนื่อย"]
                        gsheet_df["สถิติโดนยอด"] = gsheet_df.apply(
                            lambda row: row["สถิติโดนยอด"] + เหนื่อย if row["ชื่อเต็ม"] in df["ชื่อเต็ม"].values else row["สถิติโดนยอด"],
                            axis=1
                        )
            
                    # ✅ อัปเดตคอลัมน์เดียวทั้งหมดในครั้งเดียว
                    updated_column_values = gsheet_df["สถิติโดนยอด"].astype(str).tolist()
                    start_cell = 'N2'
                    end_cell = f'N{1 + len(updated_column_values)}'
                    cell_range = f'{start_cell}:{end_cell}'
                    ws.update(cell_range, [[val] for val in updated_column_values])
                    time.sleep(2)
                    st.success("✅ อัปเดต 'สถิติโดนยอด' สำเร็จ")
                    st.markdown(f"[🔗 ดูสถิติที่อัปเดตแล้ว (ชีท: {sheet_name})]({sheet_url})", unsafe_allow_html=True)
            
                except Exception as e:
                    st.error(f"❌ ไม่สามารถประมวลผลไฟล์: {e}")
        except Exception as e:
                st.error(f"❌ Error: {e}")

def ceremony_duty_mode():
    st.info("คุณเลือก: จัดยอดพิธี")

    sheet = connect_gsheet(st.session_state["sheet_name"])
    data = sheet.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])  # ข้าม header แรก

    if "สถิติโดนยอด" in df.columns:
        df["สถิติโดนยอด"] = pd.to_numeric(df["สถิติโดนยอด"], errors="coerce").fillna(0)

    ยอด_name = st.text_input("🔖กรอกชื่อยอด")
    จำนวนคน = st.number_input("👥จำนวนคน", min_value=1, step=1)

    ตัวเลือก_หน้าที่ = ["ชั้นกรม", "ชั้นพัน", "ฝอ.1", "ฝอ.4", "ฝอ.5", "เเซนเฮิร์ท", "อิสลาม", "คริสต์"]
    ตัวเลือก_หน้าที่_ทั้งหมด = ["เลือกทั้งหมด"] + ตัวเลือก_หน้าที่
    ตัวกรอง_หน้าที่_เลือก = st.multiselect("⛔ไม่เลือกคนที่มีหน้าที่", ตัวเลือก_หน้าที่_ทั้งหมด)
    
    ตัวเลือก_ชมรม = ["กรีฑา", "จักรยาน", "ไซเบอร์", "ดนตรีไทย", "ดนตรีสากล", "ดาบสากล", "นิเทศ", "สตส", "บาส", "โปโลน้ำ", "ฟุตบอล", "ยูโด", "รักบี้", "แบตมินตัน"]
    ตัวเลือก_ชมรม_ทั้งหมด = ["เลือกทั้งหมด"] + ตัวเลือก_ชมรม
    excluded_clubs = st.multiselect("⛔ไม่เลือกชมรม", ตัวเลือก_ชมรม_ทั้งหมด)
    
    if st.button("📤 จัดยอดและส่งออกไฟล์"):
        # จัดการเงื่อนไข "เลือกทั้งหมด"
        ตัวกรอง_หน้าที่ = ตัวเลือก_หน้าที่ if "เลือกทั้งหมด" in ตัวกรอง_หน้าที่_เลือก else ตัวกรอง_หน้าที่_เลือก
        filtered_clubs = ตัวเลือก_ชมรม if "เลือกทั้งหมด" in excluded_clubs else excluded_clubs
    
        df_filtered = df.copy()
        if "หน้าที่" in df_filtered.columns and ตัวกรอง_หน้าที่:
            df_filtered = df_filtered[~df_filtered["หน้าที่"].isin(ตัวกรอง_หน้าที่)]
        if "ชมรม" in df_filtered.columns and filtered_clubs:
            df_filtered = df_filtered[~df_filtered["ชมรม"].isin(filtered_clubs)]
        if "สถิติโดนยอด" in df_filtered.columns:
            df_filtered = df_filtered.sort_values(by="สถิติโดนยอด", ascending=True)
    
        grouped = df_filtered.groupby("สังกัด")
        สังกัด_list = list(grouped.groups.keys())
        คนต่อสังกัด = defaultdict(list)
    
        while sum(len(v) for v in คนต่อสังกัด.values()) < จำนวนคน:
            for สังกัด in สังกัด_list:
                available = grouped.get_group(สังกัด)
                used_indices = set().union(*คนต่อสังกัด.values())
                choices = available[~available.index.isin(used_indices)]
                if not choices.empty and sum(len(v) for v in คนต่อสังกัด.values()) < จำนวนคน:
                    chosen = choices.sample(1)
                    คนต่อสังกัด[สังกัด].append(chosen.index[0])
    
        selected_indices = [i for indices in คนต่อสังกัด.values() for i in indices]
        selected_df = df.loc[selected_indices]
        selected_df = selected_df.reset_index(drop=True)
        selected_df.index += 1
    
        if "ลำดับ" in selected_df.columns:
            selected_df = selected_df.drop(columns=["ลำดับ"])
        selected_df.insert(0, "ลำดับ", selected_df.index)
    
        selected_df["ยศ"] = "นนร."
        selected_df["ชื่อ"] = selected_df.iloc[:, 2].fillna("")
        selected_df["สกุล"] = selected_df.iloc[:, 3].fillna("")
        selected_df["ยศ ชื่อ-สกุล"] = selected_df["ยศ"] + " " + selected_df["ชื่อ"] + " " + selected_df["สกุล"]
    
        # แสดงตารางบนหน้าเว็บพร้อมคอลัมน์เบอร์โทรศัพท์
        columns = ["ลำดับ", "ยศ ชื่อ-สกุล", "ชั้นปีที่", "ตอน", "ตำแหน่ง", "สังกัด", "เบอร์โทรศัพท์", "หมายเหตุ"]
        output_df = selected_df[columns]
    
        def render_centered_table(df):
            html = """
            <style>
                table.custom-table {
                    width: 100%;
                    border-collapse: collapse;
                    table-layout: auto;
                    font-size: 11px;
                    overflow-x: auto;
                }
                table.custom-table th, table.custom-table td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: center;
                    height: 40px;
                }
                table.custom-table th {
                    font-weight: bold;
                }
                table.custom-table th:nth-child(1), table.custom-table td:nth-child(1) { width: 5%; }
                table.custom-table th:nth-child(2), table.custom-table td:nth-child(2) { width: 20%; }
                table.custom-table th:nth-child(3), table.custom-table td:nth-child(3) { width: 8%; }
                table.custom-table th:nth-child(4), table.custom-table td:nth-child(4) { width: 5%; }
                table.custom-table th:nth-child(5), table.custom-table td:nth-child(5) { width: 15%; }
                table.custom-table th:nth-child(6), table.custom-table td:nth-child(6) { width: 15%; }
                table.custom-table th:nth-child(7), table.custom-table td:nth-child(7) { width: 15%; }
                table.custom-table th:nth-child(8), table.custom-table td:nth-child(8) { width: 10%; }
                table.custom-table td:nth-child(2) {
                    text-align: left;
                    padding-left: 10px;
                }
                @media (max-width: 600px) {
                    table.custom-table {
                        font-size: 9px;
                        min-width: 600px;
                        overflow-x: auto;
                        display: block;
                    }
                }
            </style>
            """
            html += "<div style='overflow-x:auto;'>"
            html += "<table class='custom-table'>"
            html += "<thead><tr>" + "".join(f"<th>{col}</th>" for col in df.columns) + "</tr></thead>"
            html += "<tbody>"
            for _, row in df.iterrows():
                html += "<tr>"
                for i, cell in enumerate(row):
                    value = "" if pd.isna(cell) and i == 7 else cell
                    html += f"<td>{value}</td>"
                html += "</tr>"
            html += "</tbody></table>"
            html += "</div>"
            st.markdown(html, unsafe_allow_html=True)
    
        render_centered_table(output_df)
    
        # สร้าง Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "ยอดพิธี"
        ws.append([ยอด_name])
        ws.append([])
        ws.merge_cells('A2:j2')
    
        selected_df["ยศ"] = "นนร."
        selected_df["ชื่อ"] = selected_df.iloc[:, 2]
        selected_df["สกุล"] = selected_df.iloc[:, 3]
    
        columns_excel = ["ลำดับ", "ยศ", "ชื่อ", "สกุล", "ชั้นปีที่", "ตอน", "ตำแหน่ง", "สังกัด", "เบอร์โทรศัพท์", "หมายเหตุ"]
        output_df_excel = selected_df[columns_excel]
    
        # เขียนหัวตาราง Excel
        ws.append(["ลำดับ", "ยศ", "ชื่อ", "สกุล", "ชั้นปีที่", "ตอน", "ตำแหน่ง", "สังกัด", "เบอร์โทรศัพท์", "หมายเหตุ"])
        ws.merge_cells('A1:J1')
        ws.merge_cells('A2:J2')
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
        ws.cell(row=3, column=2).value = "ยศ ชื่อ-สกุล"
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=3, column=5).value = "ชั้นปีที่"
        ws.cell(row=3, column=6).value = "ตอน"
        ws.cell(row=3, column=7).value = "ตำแหน่ง"
        ws.cell(row=3, column=8).value = "สังกัด"
        ws.cell(row=3, column=9).value = "เบอร์โทรศัพท์"
        ws.cell(row=3, column=10).value = "หมายเหตุ"
    
        for r in dataframe_to_rows(output_df_excel, index=False, header=False):
            ws.append(r)
    
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row[:10]):
                if idx < 1 or idx > 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
    
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
    
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
        output_filename = f"{ยอด_name}.xlsx"
        wb.save(output_filename)
        st.success(f"✅ สร้างไฟล์ Excel สำเร็จ: {output_filename}")
        with open(output_filename, "rb") as f:
            st.download_button("📥 ดาวน์โหลด Excel", f, file_name=output_filename)


# =================== MAIN ===================
render_header()
render_menu()

mode = st.session_state.get("mode", None)

if mode == "night_duty":
    night_duty_mode()
elif mode == "weekend_duty":
    weekend_duty_mode()
elif mode == "home":
    home_mode()
elif mode == "count":
    count_mode()
elif mode == "ceremony_duty":
    ceremony_duty_mode()

st.markdown("<hr style='border:0.5px solid #ccc;'>", unsafe_allow_html=True)
st.markdown("<p style='text-align: center;'>J.A.R.V.I.S © 2025 | Dev by Oat</p>", unsafe_allow_html=True)
